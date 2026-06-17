# excel_handler.py - Đọc/ghi dữ liệu từ file Excel (.xlsx)
# Sử dụng thư viện openpyxl

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from models import SinhVien, MonHoc, BangDiem


def import_sinh_vien(filepath):
    """
    Đọc danh sách sinh viên từ file Excel.
    
    Cấu trúc file Excel (Sheet đầu tiên):
    Hàng 1: Header (MSSV | Họ tên | Ngày sinh | Giới tính | Lớp | Email)
    Hàng 2+: Dữ liệu
    
    Args:
        filepath: Đường dẫn file .xlsx
        
    Returns:
        tuple: (list[SinhVien], list[str]) - (danh sách SV, danh sách lỗi)
    """
    ds_sv = []
    errors = []
    
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active  # Lấy sheet đầu tiên
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row is None or all(cell is None for cell in row):
                continue  # Bỏ qua hàng trống
            
            try:
                # Đọc các cột, xử lý None
                mssv = str(row[0]).strip() if row[0] is not None else ""
                ho_ten = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
                ngay_sinh = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
                gioi_tinh = str(row[3]).strip() if len(row) > 3 and row[3] is not None else "Nam"
                lop = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ""
                email = str(row[5]).strip() if len(row) > 5 and row[5] is not None else ""
                
                if not mssv or mssv == "None":
                    continue
                if not ho_ten or ho_ten == "None":
                    errors.append(f"Hàng {row_idx}: Thiếu họ tên cho MSSV '{mssv}'")
                    continue
                
                sv = SinhVien(mssv, ho_ten, ngay_sinh, gioi_tinh, lop, email)
                ds_sv.append(sv)
                
            except Exception as e:
                errors.append(f"Hàng {row_idx}: Lỗi đọc dữ liệu - {str(e)}")
        
        wb.close()
        
    except FileNotFoundError:
        errors.append(f"Không tìm thấy file: {filepath}")
    except Exception as e:
        errors.append(f"Lỗi đọc file Excel: {str(e)}")
    
    return ds_sv, errors


def import_mon_hoc(filepath):
    """
    Đọc danh sách môn học từ file Excel.
    
    Cấu trúc: Mã môn | Tên môn | Số tín chỉ
    
    Returns:
        tuple: (list[MonHoc], list[str])
    """
    ds_mh = []
    errors = []
    
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row is None or all(cell is None for cell in row):
                continue
            
            try:
                ma_mon = str(row[0]).strip() if row[0] is not None else ""
                ten_mon = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
                so_tc = int(row[2]) if len(row) > 2 and row[2] is not None else 0
                
                if not ma_mon or ma_mon == "None":
                    continue
                if not ten_mon or ten_mon == "None":
                    errors.append(f"Hàng {row_idx}: Thiếu tên môn cho mã '{ma_mon}'")
                    continue
                
                mh = MonHoc(ma_mon, ten_mon, so_tc)
                ds_mh.append(mh)
                
            except Exception as e:
                errors.append(f"Hàng {row_idx}: Lỗi đọc dữ liệu - {str(e)}")
        
        wb.close()
        
    except Exception as e:
        errors.append(f"Lỗi đọc file Excel: {str(e)}")
    
    return ds_mh, errors


def import_diem(filepath):
    """
    Đọc bảng điểm từ file Excel.
    
    Cấu trúc: MSSV | Mã môn | Điểm quá trình | Điểm thi
    
    Returns:
        tuple: (list[BangDiem], list[str])
    """
    ds_diem = []
    errors = []
    
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row is None or all(cell is None for cell in row):
                continue
            
            try:
                mssv = str(row[0]).strip() if row[0] is not None else ""
                ma_mon = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
                diem_qt = float(row[2]) if len(row) > 2 and row[2] is not None else 0.0
                diem_thi = float(row[3]) if len(row) > 3 and row[3] is not None else 0.0
                
                if not mssv or mssv == "None":
                    continue
                if not ma_mon or ma_mon == "None":
                    errors.append(f"Hàng {row_idx}: Thiếu mã môn cho MSSV '{mssv}'")
                    continue
                
                # Validate điểm
                if diem_qt < 0 or diem_qt > 10:
                    errors.append(f"Hàng {row_idx}: Điểm QT không hợp lệ ({diem_qt})")
                    continue
                if diem_thi < 0 or diem_thi > 10:
                    errors.append(f"Hàng {row_idx}: Điểm thi không hợp lệ ({diem_thi})")
                    continue
                
                bd = BangDiem(mssv, ma_mon, round(diem_qt, 2), round(diem_thi, 2))
                ds_diem.append(bd)
                
            except Exception as e:
                errors.append(f"Hàng {row_idx}: Lỗi đọc dữ liệu - {str(e)}")
        
        wb.close()
        
    except Exception as e:
        errors.append(f"Lỗi đọc file Excel: {str(e)}")
    
    return ds_diem, errors


def export_template_sinh_vien(filepath):
    """Tạo file Excel mẫu cho danh sách sinh viên."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Danh sách Sinh viên"
    
    # Header style
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    
    headers = ["MSSV", "Họ tên", "Ngày sinh", "Giới tính", "Lớp", "Email"]
    col_widths = [12, 25, 15, 12, 15, 30]
    
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    
    # Dữ liệu mẫu
    sample_data = [
        ["SV001", "Nguyễn Văn An", "15/03/2003", "Nam", "CNTT01", "an.nv@email.com"],
        ["SV002", "Trần Thị Bình", "22/07/2003", "Nữ", "CNTT01", "binh.tt@email.com"],
        ["SV003", "Lê Hoàng Cường", "10/01/2004", "Nam", "CNTT02", "cuong.lh@email.com"],
        ["SV004", "Phạm Minh Duy", "05/11/2003", "Nam", "CNTT02", "duy.pm@email.com"],
        ["SV005", "Hoàng Thị Em", "30/09/2003", "Nữ", "CNTT01", "em.ht@email.com"],
    ]
    
    data_align = Alignment(horizontal="center", vertical="center")
    
    for row_idx, data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx != 2 and col_idx != 6:  # Không căn giữa họ tên và email
                cell.alignment = data_align
    
    wb.save(filepath)
    wb.close()


def export_template_mon_hoc(filepath):
    """Tạo file Excel mẫu cho danh sách môn học."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Danh sách Môn học"
    
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    
    headers = ["Mã môn", "Tên môn", "Số tín chỉ"]
    col_widths = [15, 35, 15]
    
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    
    sample_data = [
        ["CS101", "Nhập môn lập trình", 3],
        ["CS102", "Cấu trúc dữ liệu và giải thuật", 4],
        ["CS103", "Cơ sở dữ liệu", 3],
        ["CS104", "Mạng máy tính", 3],
        ["MATH01", "Toán cao cấp", 4],
    ]
    
    data_align = Alignment(horizontal="center", vertical="center")
    
    for row_idx, data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx != 2:
                cell.alignment = data_align
    
    wb.save(filepath)
    wb.close()


def export_template_diem(filepath):
    """Tạo file Excel mẫu cho bảng điểm."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bảng điểm"
    
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    
    headers = ["MSSV", "Mã môn", "Điểm quá trình", "Điểm thi"]
    col_widths = [12, 12, 18, 12]
    
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    
    sample_data = [
        ["SV001", "CS101", 8.5, 7.0],
        ["SV001", "CS102", 7.0, 8.5],
        ["SV002", "CS101", 9.0, 8.0],
        ["SV002", "CS102", 6.5, 7.5],
        ["SV003", "CS101", 8.0, 9.0],
    ]
    
    data_align = Alignment(horizontal="center", vertical="center")
    
    for row_idx, data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = data_align
    
    wb.save(filepath)
    wb.close()
